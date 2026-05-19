import json
import math
import re
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook


RETRIEVED_AT = "2026-05-19"
SOURCE_ID = "london-development-database-archive"
CORPUS_PATH = Path("data/manual_drops/architecture_milestones/architecture_milestones_2008_2026.json")
INPUT_DIR = Path("tmp/subagents/round124_london_ldd_archive_completions")
INPUT_XLSX = INPUT_DIR / "LDD_Permissions_for_Datastore_final.xlsx"
OUT_DIR = Path("tmp/subagents/round129_london_ldd_archive_more")
OUT_PATH = OUT_DIR / "candidates.json"

DATASET_PAGE = "https://data.london.gov.uk/dataset/planning-permissions-on-the-london-development-database-ldd-2jxq0/"
PERMISSIONS_DOWNLOAD = "https://data.london.gov.uk/download/2jxq0/eb050c40-3e94-4384-8e59-1b8c49dbdf36/LDD%20Permissions%20for%20Datastore%20final.xlsx"

PUBLIC_RE = re.compile(
    r"\b("
    r"school|academy|college|university|campus|hospital|health|clinic|surgery|care home|library|museum|gallery|"
    r"theatre|cinema|arts?|cultural|community|leisure|sports?|stadium|pool|church|mosque|synagogue|temple|"
    r"police|fire station|court|civic|town hall|council|station|transport|terminal|market|hotel|hostel|"
    r"public realm|square|park|playground|regeneration|estate regeneration|mixed use|mixed-use|town centre|"
    r"retail|shopping|office|commercial|employment|industrial|warehouse|workshop|laboratory|research"
    r")\b",
    re.IGNORECASE,
)

SMALL_DOMESTIC_RE = re.compile(
    r"\b("
    r"single[- ]storey extension|rear extension|loft conversion|porch|conservatory|garage conversion|"
    r"one bed house|two bed house|single dwelling|existing dwelling"
    r")\b",
    re.IGNORECASE,
)


def clean_text(value):
    if value is None:
        return ""
    text = (
        str(value)
        .replace("_x000D_", " ")
        .replace("\u2019", "'")
        .replace("\u2013", "-")
        .replace("\u2014", "-")
    )
    return re.sub(r"\s+", " ", text).strip()


def slugify(value, limit=96):
    text = clean_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text[:limit].rstrip("_") or "ldd_record"


def as_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean_text(value)
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], fmt).date().isoformat()
        except ValueError:
            pass
    return ""


def number(value):
    try:
        if value is None or value == "":
            return 0.0
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def osgb36_to_wgs84(easting, northing):
    # Airy 1830 ellipsoid, British National Grid projection, then Helmert transform to WGS84.
    a = 6377563.396
    b = 6356256.909
    f0 = 0.9996012717
    lat0 = math.radians(49)
    lon0 = math.radians(-2)
    n0 = -100000
    e0 = 400000
    e2 = 1 - (b * b) / (a * a)
    n = (a - b) / (a + b)

    lat = lat0
    m = 0
    while northing - n0 - m >= 0.00001:
        lat = (northing - n0 - m) / (a * f0) + lat
        ma = (1 + n + (5 / 4) * n**2 + (5 / 4) * n**3) * (lat - lat0)
        mb = (3 * n + 3 * n**2 + (21 / 8) * n**3) * math.sin(lat - lat0) * math.cos(lat + lat0)
        mc = ((15 / 8) * n**2 + (15 / 8) * n**3) * math.sin(2 * (lat - lat0)) * math.cos(2 * (lat + lat0))
        md = (35 / 24) * n**3 * math.sin(3 * (lat - lat0)) * math.cos(3 * (lat + lat0))
        m = b * f0 * (ma - mb + mc - md)

    sin_lat = math.sin(lat)
    cos_lat = math.cos(lat)
    nu = a * f0 / math.sqrt(1 - e2 * sin_lat**2)
    rho = a * f0 * (1 - e2) / (1 - e2 * sin_lat**2) ** 1.5
    eta2 = nu / rho - 1
    tan_lat = math.tan(lat)
    sec_lat = 1 / cos_lat
    d_e = easting - e0

    vii = tan_lat / (2 * rho * nu)
    viii = tan_lat / (24 * rho * nu**3) * (5 + 3 * tan_lat**2 + eta2 - 9 * tan_lat**2 * eta2)
    ix = tan_lat / (720 * rho * nu**5) * (61 + 90 * tan_lat**2 + 45 * tan_lat**4)
    x = sec_lat / nu
    xi = sec_lat / (6 * nu**3) * (nu / rho + 2 * tan_lat**2)
    xii = sec_lat / (120 * nu**5) * (5 + 28 * tan_lat**2 + 24 * tan_lat**4)
    xiia = sec_lat / (5040 * nu**7) * (61 + 662 * tan_lat**2 + 1320 * tan_lat**4 + 720 * tan_lat**6)

    lat_osgb = lat - vii * d_e**2 + viii * d_e**4 - ix * d_e**6
    lon_osgb = lon0 + x * d_e - xi * d_e**3 + xii * d_e**5 - xiia * d_e**7

    h = 0
    sin_phi = math.sin(lat_osgb)
    cos_phi = math.cos(lat_osgb)
    sin_lam = math.sin(lon_osgb)
    cos_lam = math.cos(lon_osgb)
    nu = a / math.sqrt(1 - e2 * sin_phi**2)
    x1 = (nu + h) * cos_phi * cos_lam
    y1 = (nu + h) * cos_phi * sin_lam
    z1 = ((1 - e2) * nu + h) * sin_phi

    tx, ty, tz = 446.448, -125.157, 542.060
    s = 20.4894 * 1e-6
    rx = math.radians(0.1502 / 3600)
    ry = math.radians(0.2470 / 3600)
    rz = math.radians(0.8421 / 3600)
    x2 = tx + (1 + s) * x1 + (-rz) * y1 + ry * z1
    y2 = ty + rz * x1 + (1 + s) * y1 + (-rx) * z1
    z2 = tz + (-ry) * x1 + rx * y1 + (1 + s) * z1

    a2 = 6378137.000
    b2 = 6356752.3141
    e22 = 1 - (b2 * b2) / (a2 * a2)
    p = math.sqrt(x2 * x2 + y2 * y2)
    lat2 = math.atan2(z2, p * (1 - e22))
    while True:
        nu2 = a2 / math.sqrt(1 - e22 * math.sin(lat2) ** 2)
        next_lat = math.atan2(z2 + e22 * nu2 * math.sin(lat2), p)
        if abs(next_lat - lat2) < 1e-12:
            lat2 = next_lat
            break
        lat2 = next_lat
    lon2 = math.atan2(y2, x2)
    return round(math.degrees(lat2), 6), round(math.degrees(lon2), 6)


def existing_ldd_rows():
    doc = json.loads(CORPUS_PATH.read_text(encoding="utf-8-sig"))
    rows = set()
    refs = set()
    for event in doc.get("events", []):
        if SOURCE_ID not in event.get("source_ids", []):
            continue
        record_id = clean_text(event.get("source_record_id"))
        match = re.search(r"workbook row (\d+)", record_id)
        if match:
            rows.add(int(match.group(1)))
        ref = re.search(r"planning_authority=([^;]+);\s*borough_reference=(.+)$", record_id)
        if ref:
            refs.add((clean_text(ref.group(1)).lower(), clean_text(ref.group(2)).lower()))
    return rows, refs


def score_row(row):
    text = " ".join(clean_text(row.get(k)) for k in (
        "Development Description",
        "Scheme Name",
        "Site Name/Number",
        "Primary Street Name",
        "Permission Type",
    ))
    residential_units = number(row.get("Proposed Total Residential Units"))
    affordable_units = number(row.get("Proposed Total Affordable Units"))
    floorspace = number(row.get("Proposed Total Floorspace"))
    site_area = number(row.get("Total Site Area (Proposed) Hectares (ha)"))
    non_res_area = number(row.get("Non Res Site Area (Proposed) Hectares (ha)"))

    score = 0
    if PUBLIC_RE.search(text):
        score += 25
    if residential_units >= 100:
        score += min(30, residential_units / 40)
    if affordable_units >= 20:
        score += min(20, affordable_units / 20)
    if floorspace >= 5000:
        score += min(25, floorspace / 4000)
    if site_area >= 1:
        score += min(20, site_area * 2)
    if non_res_area >= 0.5:
        score += min(10, non_res_area * 2)
    if re.search(r"\b(outline|reserved matters|variation|section 73|details|discharge)\b", text, re.I):
        score -= 3
    if SMALL_DOMESTIC_RE.search(text) and residential_units < 20 and floorspace < 2000:
        score -= 30
    return score


def label_for(row):
    scheme = clean_text(row.get("Scheme Name"))
    site = clean_text(row.get("Site Name/Number"))
    street = clean_text(row.get("Primary Street Name"))
    if scheme:
        return scheme
    if site and street:
        return f"{site}, {street}"
    return site or street or clean_text(row.get("Development Description"))[:80] or "LDD completion record"


def candidate_for(row, excel_row):
    completed_date = as_date(row.get("Date construction completed (Completed Date)"))
    permission_date = as_date(row.get("Permission Date"))
    started_date = as_date(row.get("Date work commenced on site (Started Date)"))
    easting = number(row.get("Easting"))
    northing = number(row.get("Northing"))
    lat, lon = osgb36_to_wgs84(easting, northing)

    authority = clean_text(row.get("Planning Authority"))
    borough_ref = clean_text(row.get("Borough Reference"))
    title_label = label_for(row)
    description = clean_text(row.get("Development Description"))
    residential_units = int(number(row.get("Proposed Total Residential Units")))
    affordable_units = int(number(row.get("Proposed Total Affordable Units")))
    floorspace = int(number(row.get("Proposed Total Floorspace")))
    site_area = number(row.get("Total Site Area (Proposed) Hectares (ha)"))

    metrics = []
    if residential_units:
        metrics.append(f"{residential_units} proposed homes")
    if affordable_units:
        metrics.append(f"{affordable_units} proposed affordable homes")
    if floorspace:
        metrics.append(f"{floorspace} sqm proposed floorspace")
    if site_area:
        metrics.append(f"{site_area:g} ha proposed site area")
    metric_text = "; ".join(metrics) or "no headline quantity fields supplied"

    source_record_id = f"LDD planning permissions workbook row {excel_row}; planning_authority={authority}; borough_reference={borough_ref}"

    return {
        "city_id": "london",
        "candidate_id": f"lon_ldd_archive_completion_round129_row_{excel_row}_{slugify(authority)}_{slugify(borough_ref, 48)}",
        "date": completed_date,
        "date_precision": "day",
        "bucket": "planning/development/architecture/ldd_completion_record",
        "title": f"LDD completion record: {title_label}",
        "summary": f"The London Development Database archive records {title_label} in {authority} as Completed on {completed_date}. The row records {metric_text}. Source proposal description: {description}",
        "observed_change": "LDD archived planning-permission row records the permission status as Completed and supplies a construction-completed date.",
        "area": f"{authority}; {title_label}",
        "latitude": lat,
        "longitude": lon,
        "source_ids": [SOURCE_ID],
        "source_name": "Planning permissions on the London Development Database (LDD)",
        "publisher": "Greater London Authority / London Development Database / relevant London planning authority",
        "source_url": DATASET_PAGE,
        "source_record_id": source_record_id,
        "source_type": "official archived spreadsheet row",
        "accessed_at": RETRIEVED_AT,
        "source_date_field": "Date construction completed (Completed Date)",
        "source_dataset_id": SOURCE_ID,
        "confidence": "documented",
        "architect": "Source record does not name a project architect.",
        "project_type": "LDD archived development completion-status record",
        "geometry_source": "LDD Easting/Northing point fields from the official planning-permissions workbook, converted from British National Grid to WGS84.",
        "geometry_precision": "LDD point location only; not a building footprint, parcel boundary, entrance, or current as-built geometry.",
        "license_or_terms_note": "London Datastore / Greater London Authority archive under the UK Open Government Licence v3.0; preserve workbook row references and attribution.",
        "attribution": "Greater London Authority / London Development Database / relevant London planning authority",
        "limitations": "This is an archived LDD planning/development completion-status record. Completed Date has no formal single definition in the LDD notes and may reflect a building-control completion certificate or planning-authority judgement. It is not evidence of public opening, occupation, current use, service delivery, final built form, design quality, local outcomes, or causal relationships. Some rows may represent phases, outline permissions, reserved matters, variations, or administrative updates.",
        "permission_date": permission_date,
        "started_date": started_date,
        "completed_financial_year": clean_text(row.get("Completed Financial Year")),
        "planning_authority": authority,
        "borough_reference": borough_ref,
        "permission_type": clean_text(row.get("Permission Type")),
        "decision_agency": clean_text(row.get("Decision Agency")),
        "current_permission_status": clean_text(row.get("Current permission status")),
        "source_file_url": PERMISSIONS_DOWNLOAD,
        "source_row_references": [
            {
                "workbook": "LDD_Permissions_for_Datastore_final.xlsx",
                "sheet": "LDD data",
                "excel_row": excel_row,
                "row_reference": f"{authority} / {borough_ref}",
                "download_url": PERMISSIONS_DOWNLOAD,
            }
        ],
        "raw_row": {
            "planning_authority": authority,
            "borough_reference": borough_ref,
            "development_description": description,
            "scheme_name": clean_text(row.get("Scheme Name")),
            "site_name_number": clean_text(row.get("Site Name/Number")),
            "primary_street_name": clean_text(row.get("Primary Street Name")),
            "postcode": clean_text(row.get("Post Code")),
            "ward": clean_text(row.get("Ward")),
            "easting": easting,
            "northing": northing,
            "permission_date": permission_date,
            "started_date": started_date,
            "completed_date": completed_date,
            "proposed_total_residential_units": residential_units,
            "proposed_total_affordable_units": affordable_units,
            "proposed_total_floorspace": floorspace,
            "total_site_area_ha": site_area,
        },
    }


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    existing_rows, existing_refs = existing_ldd_rows()

    workbook = load_workbook(INPUT_XLSX, read_only=True, data_only=True)
    worksheet = workbook["LDD data"]
    headers = list(next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True)))

    scored = []
    rejected = []
    for excel_row, values in enumerate(worksheet.iter_rows(min_row=3, values_only=True), start=3):
        row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        authority = clean_text(row.get("Planning Authority"))
        borough_ref = clean_text(row.get("Borough Reference"))
        completed_date = as_date(row.get("Date construction completed (Completed Date)"))
        easting = number(row.get("Easting"))
        northing = number(row.get("Northing"))
        if excel_row in existing_rows or (authority.lower(), borough_ref.lower()) in existing_refs:
            continue
        if clean_text(row.get("Current permission status")).lower() != "completed":
            continue
        if not ("2008-01-01" <= completed_date <= RETRIEVED_AT):
            continue
        if not (500000 <= easting <= 565000 and 155000 <= northing <= 205000):
            rejected.append({"excel_row": excel_row, "reason": "Missing or outside-London LDD point.", "planning_authority": authority, "borough_reference": borough_ref})
            continue
        score = score_row(row)
        if score < 22:
            continue
        scored.append((score, completed_date, excel_row, row))

    scored.sort(key=lambda item: (-item[0], item[1], item[2]))
    candidates = [candidate_for(row, excel_row) for score, completed_date, excel_row, row in scored[:220]]

    payload = {
        "generated_at": RETRIEVED_AT,
        "source_audits": [
            {
                "source_id": SOURCE_ID,
                "source_name": "Planning permissions on the London Development Database (LDD)",
                "publisher": "Greater London Authority / London Development Database / relevant London planning authority",
                "source_url": DATASET_PAGE,
                "api_endpoint": PERMISSIONS_DOWNLOAD,
                "license_or_terms_note": "London Datastore / Greater London Authority archive under the UK Open Government Licence v3.0.",
                "coverage_years_checked": "Completed LDD planning-permissions rows from 2008-01-01 through 2026-05-19; archived workbooks report starts/completions up to 2019/2020 and are no longer updated after PLD replacement.",
                "update_frequency": "One-off archived Datastore workbook; no longer updated.",
                "geographic_scope": "Greater London LDD point rows from planning authorities.",
                "key_fields_used": "Planning Authority, Borough Reference, Current permission status, Development Description, Scheme Name, Easting, Northing, Permission Date, Started Date, Completed Date, proposed units/floorspace/site area.",
                "reliability": "strong for archived LDD administrative completion-status records; usable with caveats for city-change events",
                "ingestion_recommendation": "Use only as LDD completion-status milestones with visible caveats; do not infer opening, occupation, final built form, current use, outcomes, or causation.",
            }
        ],
        "selection_summary": {
            "input_workbook": str(INPUT_XLSX),
            "eligible_scored_rows": len(scored),
            "retained_candidates": len(candidates),
            "excluded_existing_ldd_rows": len(existing_rows),
            "candidate_limit": 220,
        },
        "candidates": candidates,
        "rejected": rejected[:250],
    }

    OUT_PATH.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "eligibleScoredRows": len(scored),
        "candidates": len(candidates),
        "existingLddRows": len(existing_rows),
        "outPath": str(OUT_PATH),
    }, indent=2))


if __name__ == "__main__":
    main()
